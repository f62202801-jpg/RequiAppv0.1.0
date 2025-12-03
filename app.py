# app.py 
# Single-file Flask application with:
# - Requisitions (create)
# - Mis requisiciones (view)
# - Autorizaciones (jefes y admin)
# - Admin: manage catalog, upload images, download & clear requisitions (saved to static/descargas)
# - Projects can be added as "proyecto especial" and persisted to projects.json

from flask import Flask, render_template, request, redirect, session, url_for, send_from_directory, jsonify, flash, send_file
from flask_cors import CORS
import os, json, datetime, uuid, smtplib, csv
from email.message import EmailMessage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from werkzeug.utils import secure_filename
import pandas as pd

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get("APP_SECRET", "CAMBIAR_POR_SECRETO_PROD")

BASE_DIR = os.path.dirname(__file__)
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

CATALOG_FILE = os.path.join(BASE_DIR, "Control De Herramientas.xlsm")
CATALOG_JSON = os.path.join(BASE_DIR, "catalog.json")
REQUISITIONS_FILE = os.path.join(BASE_DIR, "Requisiciones.xlsx")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
DOWNLOADS_FOLDER = os.path.join(BASE_DIR, "static", "descargas")
PROJECTS_JSON = os.path.join(BASE_DIR, "projects.json")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOADS_FOLDER, exist_ok=True)

ALLOWED_EXT = {"png", "jpg", "jpeg"}

# --- Configuración correo (opcional) ---
SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587) or 587)
SMTP_USER = os.environ.get("SMTP_USER", "eoronzor@latinberryplants.com")
SMTP_PASS = os.environ.get("SMTP_PASS", "20940012(EkO)")
NOTIFY_EMAILS = os.environ.get("NOTIFY_EMAILS", "eoronzor@latinberryplants.com")

# --- Usuarios ---
USERS = {
    
    "carlos": {
        "password": "123456",
        "role": "comprador",
        "name": "Carlos Montiel",
        "jefe": "Diana"
    },
    "erik": {
        "password": "1234",
        "role": "solicitante",
        "name": "Erik Oronzor",
        "jefe": "Alfonso"
    },
    "juanjo": {
        "password": "juan123",
        "role": "solicitante",
        "name": "Juan José M.",
        "jefe": "iliana"
    },
    "Emid": {
        "password": "123",
        "role": "solicitante",
        "name": "Emid Tellez",
        "jefe": "marcosb"   
    },
    "jesus": {
        "password": "admin",
        "role": "admin",
        "name": "José Jesús López",
        "jefe":"Diana"
    },
    "Yoalli": {
        "password": "1212",
        "role": "solicitante",
        "name": "Yoalli Gonzalez",
        "jefe": "Diana"
    },
    "Iliana": {
        "password": "2324",
        "role": "jefe",
        "name": "Iliana Orozco"
    },
    # --- Jefes inmediatos/Aprovadores ---
    "marcosb": {
        "password": "12345",
        "role": "jefe",
        "name": "Marcos Barrera"
    },
    "Iliana": {
        "password": "123",
        "role": "jefe",
        "name": "Iliana Orozco"
    },
    "Alfonso": {
        "password": "2025",
        "role": "jefe",
        "name": "Alfonso Aquino"
    },
    "Rebecca": {
        "password":"jefe",
        "role":"jefe",
        "name":"Rebecca Prieto"
    },
    "Diana": {
        "password":"2020",
        "role":"jefe",
        "name":"Diana López"
    }
}


DEPARTMENTS = ["Fresa", "Invernadero", "Administracion"]
STAGES = {
    "Fresa": ["Preplantacion", "Plantacion", "Crecimiento", "Cumplimiento de pedido", "Cosecha", "Postcosecha", "Mercancias en transito", "Capex"],
    "Invernadero": ["Preplantacion", "Plantacion", "Crecimiento", "Cumplimiento de pedido", "Cosecha/Postcosecha", "Mercancias en transito", "Capex"],
    "Administracion": ["Gastos de administracion", "Opex", "Capex"]
}
# Default projects_by_dept (can be extended with projects.json)
PROJECTS_BY_DEPT = {
    "Fresa": ["101 - Strawberry"],
    "Invernadero": ["103 - Raspberry Increace", "106 - Raspberry Puebla","110 - Blackberry Increce", "111 - Blueberry Puebla","207 - Blackberry Puebla", "210 - Crop Rotation Oats", "211 - Crop Rotation Barley", "212 - TNP Plantfor", "213 - TNP NAP", "216 - Raspberry Queretaro", "217 - Blackberry Queretaro", "218 - Blueberry Queretaro"],
    "Administracion": ["300 - Board Of Directors, President & Special Projects", "301 - Sales", "302 - Outdoor T&E Production", "303 - Indoor T&E Production", "304 - Finance & Controlling", "305 - HR , SSHE and Facility","306 - Procurement & Warehouse","307- Legal", "308 - IT"]
}


UNITS = ["Bulto","Gramos","Pieza", "Caja", "Metro","Mililitro", "Litro", "Kilogramo", "Tonelada", "Paquete","Par","Rollo", "Juego", "Tonelada", "Servicio"]

LOCATIONS =["Cowork Puebla", "Cowork Guadalajara","Ocotepec", "Praderas", "Purisima de Cubos","San Martin", "San Roque","Santa Julia 1","Santa Julia 2","Santa Julia 3","Santa Julia 4","Santa Julia 5","Santa Julia 6","Santa Lugarda",]


# --- Inicialización de archivos ---
def init_files():
    if not os.path.exists(REQUISITIONS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requisiciones"
        ws.append(["ID", "Fecha", "Usuario", "Solicitante", "Departamento", "Etapa", "Proyecto",
                   "Producto", "Unidad", "Cantidad", "Ubicacion", "Motivo", "Jefe",
                   "Estatus", "Fecha Autorizacion", "Comentarios Jefe", "ImageFile"])
        ws2 = wb.create_sheet("Images")
        ws2.append(["ID", "ImageFile"])
        wb.save(REQUISITIONS_FILE)
    # ensure projects.json exists
    if not os.path.exists(PROJECTS_JSON):
        with open(PROJECTS_JSON, "w", encoding="utf-8") as f:
            json.dump(PROJECTS_BY_DEPT, f, ensure_ascii=False, indent=2)

init_files()

# --- Helpers for projects persistence ---
def load_projects():
    global PROJECTS_BY_DEPT
    try:
        if os.path.exists(PROJECTS_JSON):
            with open(PROJECTS_JSON, "r", encoding="utf-8") as f:
                data = json.load(f)
            # ensure keys exist for all departments
            for d in DEPARTMENTS:
                if d not in data:
                    data[d] = PROJECTS_BY_DEPT.get(d, [])
            PROJECTS_BY_DEPT = data
    except Exception as e:
        print("Error loading projects.json:", e)

def save_projects():
    try:
        with open(PROJECTS_JSON, "w", encoding="utf-8") as f:
            json.dump(PROJECTS_BY_DEPT, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print("Error saving projects.json:", e)
        return False

load_projects()

# --- Funciones ---
def send_email(to_list, subject, body):
    if not to_list:
        return
    recipients = [e.strip() for e in to_list.split(",") if e.strip()]
    if not (SMTP_HOST and SMTP_USER and SMTP_PASS):
        print("Correo simulado a:", recipients)
        print(subject)
        return
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SMTP_USER
        msg["To"] = ", ".join(recipients)
        msg.set_content(body)
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls() 
            s.send_message(msg)
    except Exception as e:
        print("Error enviando correo:", e)

def get_products():
    if os.path.exists(CATALOG_JSON):
        try:
            with open(CATALOG_JSON,"r",encoding="utf-8") as f:
                data=json.load(f)
            return [p["name"] if isinstance(p, dict) else str(p) for p in data]
        except:
            return []
    # Fallback: leer excel
    products=[]
    if os.path.exists(CATALOG_FILE):
        try:
            wb=openpyxl.load_workbook(CATALOG_FILE,data_only=True)
            if "PRODUCTOS" in wb.sheetnames:
                ws=wb["PRODUCTOS"]
                for cell in ws['A']:
                    if cell.value:
                        products.append(str(cell.value))
        except:
            pass
    return products

def save_products(products):
    try:
        with open(CATALOG_JSON,"w",encoding="utf-8") as f:
            json.dump([{"name":p} for p in products],f,ensure_ascii=False,indent=2)
        return True
    except:
        return False

def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED_EXT

# --- Templates minimal auto-create (if missing) ---
TEMPLATE_FILES = {
    'login.html': r"""<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<title>Login</title></head>
<body class="bg-light">
<div class="container" style="max-width:420px;margin-top:8vh;">
  <div class="card p-4 shadow-sm">
    <h4 class="mb-3">Iniciar sesión</h4>
    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for cat,msg in messages %}
          <div class="alert alert-{{cat}}">{{msg}}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}
    <form method="post">
      <div class="mb-3"><label class="form-label">Usuario</label><input name="usuario" class="form-control" required></div>
      <div class="mb-3"><label class="form-label">Contraseña</label><input name="password" type="password" class="form-control" required></div>
      <div class="d-flex justify-content-between align-items-center">
        <button class="btn btn-primary">Entrar</button>
      </div>
    </form>
  </div>
</div>
</body>
</html>""",
    'solicitud.html': r"""<!doctype html>
<html lang="es">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Nueva Requisición</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
      .product-row { background:#fff;padding:12px;border-radius:8px;margin-bottom:10px;border:1px solid #e6e9ee;}
    </style>
  </head>
  <body>
    <nav class="navbar navbar-dark bg-primary">
      <div class="container">
        <a class="navbar-brand" href="#">Requisiciones LBP & LBL</a>
        <div>
          <a class="btn btn-light btn-sm me-2" href="{{ url_for('mis_requisiciones') }}">Mis requisiciones</a>
          {% set is_jefe = False %}
          {% for k,v in USERS.items() %}
            {% if v.get('jefe') == session.get('user') %}
              {% set is_jefe = True %}
            {% endif %}
          {% endfor %}
          {% if session.get('role') == 'admin' or is_jefe %}
            <a class="btn btn-light btn-sm me-2" href="{{ url_for('autorizaciones') }}">Autorizaciones</a>
          {% endif %}
          {% if session.get('role')=='admin' %}<a class="btn btn-light btn-sm me-2" href="{{ url_for('admin_products') }}">Admin</a>{% endif %}
          <a class="btn btn-danger btn-sm" href="{{ url_for('logout') }}">Salir</a>
        </div>
      </div>
    </nav>

    <div class="container mt-4">
      <div class="card p-4">
        <h4>Crear requisición</h4>
        {% with messages = get_flashed_messages(with_categories=true) %}
          {% if messages %}
            {% for cat,msg in messages %}
              <div class="alert alert-{{cat}}">{{msg}}</div>
            {% endfor %}
          {% endif %}
        {% endwith %}
        <form method="post" enctype="multipart/form-data">
          <div class="row g-3 mb-3">
            <div class="col-md-4"><label class="form-label">Departamento</label>
              <select id="departamento" name="departamento" class="form-select" required>
                {% for d in departments %}<option value="{{d}}">{{d}}</option>{% endfor %}
              </select></div>
            <div class="col-md-4"><label class="form-label">Etapa</label>
              <select id="etapa" name="etapa" class="form-select" required></select></div>
            <div class="col-md-4"><label class="form-label">Proyecto</label>
              <div class="d-flex gap-2">
                <select id="proyecto" name="proyecto" class="form-select" required style="flex:1;">
                  <!-- filled by JS -->
                </select>
                <button type="button" id="addProjectBtn" class="btn btn-outline-secondary" title="Agregar proyecto especial">+</button>
              </div>
              <input type="text" id="proyecto_especial" name="proyecto_especial" class="form-control mt-2" placeholder="Nombre de proyecto especial (opcional)" style="display:none;">
            </div>
            <div class="col-md-6"><label class="form-label">Ubicación / Rancho</label>
              <select name="ubicacion" class="form-select" required>
                {% for l in locations %}<option value="{{l}}">{{l}}</option>{% endfor %}
              </select></div>
            <div class="col-md-6"><label class="form-label">Motivo de Compras</label>
              <input name="motivo" class="form-control"></div>
          </div>

          <hr>
          <div id="productsContainer"></div>
          <div class="d-flex justify-content-between align-items-center mb-3">
            <button type="button" id="addProduct" class="btn btn-outline-primary">+ Agregar producto / servicio</button>
            <div class="legend-contact text-muted">Si el producto o servicio no aparece en el catálogo, contactar al área de Compras.</div>
          </div>

          <input type="hidden" name="products_json" id="products_json">
          <div class="text-end"><button class="btn btn-success">Enviar requisición</button></div>
        </form>
      </div>
    </div>

    <datalist id="products_list">{% for p in products %}<option value="{{ p }}"></option>{% endfor %}</datalist>

    <script>
      const stages = {{ stages|tojson }};
      const projects_by_dept = {{ projects_by_dept|tojson }};
      const units = {{ units|tojson }};

      function loadStages(dep){
        const etapaSel = document.getElementById('etapa'); etapaSel.innerHTML='';
        (stages[dep]||[]).forEach(s=>{ let o=document.createElement('option'); o.value=s;o.textContent=s; etapaSel.appendChild(o);});
      }
      function loadProjects(dep){
        const psel=document.getElementById('proyecto'); psel.innerHTML='';
        (projects_by_dept[dep]||[]).forEach(p=>{ let o=document.createElement('option'); o.value=p;o.textContent=p; psel.appendChild(o);});
        // add a special option
        let other = document.createElement('option'); other.value='__other__'; other.textContent='-- Agregar proyecto especial --'; psel.appendChild(other);
      }

      document.addEventListener('DOMContentLoaded', ()=>{
        const dep=document.getElementById('departamento').value; loadStages(dep); loadProjects(dep);
        document.getElementById('departamento').addEventListener('change',(e)=>{ 
          loadStages(e.target.value); 
          loadProjects(e.target.value); 
        });

        // Si la etapa elegida es 'Capex' => forzar proyecto en __other__ y mostrar input
        document.getElementById('etapa').addEventListener('change', function(){
          const etapaVal = (this.value || '').toLowerCase();
          if(etapaVal === 'capex' || etapaVal === 'capex' || etapaVal.includes('capex')){
            const proyectoSel = document.getElementById('proyecto');
            // ensure options exist
            if(proyectoSel.options.length === 0){
              loadProjects(document.getElementById('departamento').value);
            }
            proyectoSel.value='__other__';
            document.getElementById('proyecto_especial').style.display='block';
          }
        });

        // proyecto especial toggling
        document.getElementById('proyecto').addEventListener('change', function(){
          if(this.value==='__other__'){
            document.getElementById('proyecto_especial').style.display='block';
          } else {
            document.getElementById('proyecto_especial').style.display='none';
            document.getElementById('proyecto_especial').value='';
          }
        });
        document.getElementById('addProjectBtn').addEventListener('click', function(){
          document.getElementById('proyecto_especial').style.display = 'block';
          document.getElementById('proyecto_especial').focus();
          document.getElementById('proyecto').value='__other__';
        });

        let idx=0; function addRow(){
          const container=document.getElementById('productsContainer'); const row=document.createElement('div');
          row.className='product-row'; row.dataset.index=idx;
          row.innerHTML = `
            <div class='row g-2 align-items-end'>
              <div class='col-md-4'>
                <label class='form-label'>Producto / Servicio</label>
                <input list="products_list" class='form-control prod-input' name='product_${idx}' required placeholder="Buscar producto o servicio">
              </div>
              <div class='col-md-2'>
                <label class='form-label'>Unidad</label>
                <select class='form-select unit-select' name='unit_${idx}' required>${units.map(u=>`<option value="${u}">${u}</option>`).join('')}</select>
              </div>
              <div class='col-md-2'>
                <label class='form-label'>Cantidad</label>
                <input type='number' class='form-control qty-input' name='qty_${idx}' min="1" step="0.01" required>
              </div>
              <div class='col-md-2'>
                <label class='form-label'>Imagen</label>
                <input type='file' class='form-control img-input' name='image_${idx}' accept='image/*'>
              </div>
              <div class='col-md-2 text-end'>
                <button type='button' class='btn btn-sm btn-outline-danger remove-btn mt-4'>Eliminar</button>
              </div>
            </div>`;
          container.appendChild(row);
          row.querySelector('.remove-btn').addEventListener('click',()=>row.remove()); idx++;
        }
        addRow(); document.getElementById('addProduct').addEventListener('click',()=>addRow());

        document.querySelector('form').addEventListener('submit',(e)=>{
          const prods=[...document.querySelectorAll('.prod-input')].map(s=>s.value.trim());
          const unitsArr=[...document.querySelectorAll('.unit-select')].map(s=>s.value.trim());
          const qtys=[...document.querySelectorAll('.qty-input')].map(s=>s.value.trim());
          let lines=prods.map((p,i)=>`${p}|${unitsArr[i]}|${qtys[i]}`);
          document.getElementById('products_json').value=lines.join("\n");
        });
      });
    </script>
  </body>
</html>""",
    'mis_requisiciones.html': r"""<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<title>Mis requisiciones</title></head>
<body class="bg-light">
<div class="container mt-4">
  <h4>Mis requisiciones</h4>
  {% if requisiciones %}
  <div class="table-responsive">
    <table class="table table-sm table-bordered bg-white">
      <thead><tr>{% for h in ['ID','Fecha','Usuario','Solicitante','Departamento','Etapa','Proyecto','Producto','Unidad','Cantidad','Ubicacion','Motivo','Jefe','Estatus','Fecha Autorizacion','Comentarios Jefe','ImageFile'] %}<th>{{h}}</th>{% endfor %}</tr></thead>
      <tbody>
        {% for r in requisiciones %}
        <tr>
          {% for c in r %}<td>{{c}}</td>{% endfor %}
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% else %}
    <div class="alert alert-info">No tiene requisiciones.</div>
  {% endif %}
  <div class="mt-3"><a class="btn btn-secondary" href="{{ url_for('solicitud') }}">Regresar</a></div>
</div>
</body>
</html>""",
    'autorizaciones.html': r"""<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<title>Autorizaciones</title></head>
<body class="bg-light">
<div class="container mt-4">
  <h4>Autorizaciones pendientes</h4>
  {% if requisiciones %}
  <table class="table table-sm bg-white table-bordered">
    <thead><tr><th>ID</th><th>Solicitante</th><th>Departamento</th><th>Proyecto</th><th>Producto</th><th>Cantidad</th><th>Motivo</th><th>Acciones</th></tr></thead>
    <tbody>
      {% for r in requisiciones %}
      <tr>
        <td>{{ r[0] }}</td>
        <td>{{ r[3] }}</td>
        <td>{{ r[4] }}</td>
        <td>{{ r[6] }}</td>
        <td>{{ r[7] }}</td>
        <td>{{ r[9] }}</td>
        <td>{{ r[11] }}</td>
        <td>
          <form method="post" class="d-flex flex-column gap-2">
            <input type="hidden" name="req_id" value="{{ r[0] }}">
            <textarea name="comentario" class="form-control mb-2" placeholder="Comentario opcional"></textarea>
            <div class="btn-group">
              <button name="decision" value="aprobar" class="btn btn-success btn-sm">Aprobar</button>
              <button name="decision" value="rechazar" class="btn btn-danger btn-sm">Rechazar</button>
            </div>
          </form>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  {% else %}
    <div class="alert alert-info">No hay requisiciones pendientes.</div>
  {% endif %}

  <div class="mt-3 d-flex justify-content-between">
    <a class="btn btn-secondary" href="{{ url_for('solicitud') }}">Regresar</a>
    {% if session.get('role') == 'admin' %}
      <a class="btn btn-primary" href="{{ url_for('download_and_clear_requisitions') }}">Descargar solicitudes (y limpiar)</a>
    {% endif %}
  </div>
</div>
</body>
</html>""",
    'admin_products.html': r"""<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<title>Catálogo</title></head>
<body class="bg-light">
<div class="container mt-4">
  <h4>Administrar Catálogo</h4>
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for cat,msg in messages %}<div class="alert alert-{{cat}}">{{msg}}</div>{% endfor %}
    {% endif %}
  {% endwith %}
  <form method="post" class="mb-3">
    <input type="hidden" name="action" value="add">
    <div class="input-group"><input name="name" class="form-control" placeholder="Nombre producto/servicio"><button class="btn btn-primary">Agregar</button></div>
  </form>
  <table class="table table-sm bg-white table-bordered"><thead><tr><th>Producto</th><th>Acciones</th></tr></thead><tbody>
    {% for p in products %}
    <tr><td>{{p}}</td><td>
      <form method="post" style="display:inline-block;">
        <input type="hidden" name="action" value="delete"><input type="hidden" name="name" value="{{p}}">
        <button class="btn btn-danger btn-sm">Eliminar</button>
      </form>
    </td></tr>
    {% endfor %}
  </tbody></table>
  <div class="mt-3"><a class="btn btn-secondary" href="{{ url_for('solicitud') }}">Regresar</a></div>
</div>
</body>
</html>"""
}

# create templates if missing
for fname, content in TEMPLATE_FILES.items():
    fpath = os.path.join(TEMPLATES_DIR, fname)
    if not os.path.exists(fpath):
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(content)

# --- Rutas ---
@app.route("/", methods=["GET","POST"])
def login():
    if request.method=="POST":
        user=request.form.get("usuario")
        pwd=request.form.get("password")
        u=USERS.get(user)
        if u and u["password"]==pwd:
            session["user"]=user
            session["role"]=u["role"]
            session["name"]=u["name"]
            return redirect(url_for("solicitud"))
        flash("Usuario o contraseña incorrectos","danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/solicitud", methods=["GET","POST"])
def solicitud():
    if "user" not in session:
        return redirect(url_for("login"))
    products = get_products()
    load_projects()  # refresh projects from disk
    if request.method=="POST":
        departamento=request.form.get("departamento")
        etapa=request.form.get("etapa")
        proyecto=request.form.get("proyecto","")
        proyecto_especial = request.form.get("proyecto_especial","").strip()
        # if proyecto is __other__ or proyecto_especial provided, use proyecto_especial
        if proyecto == "__other__" and proyecto_especial:
            proyecto = proyecto_especial
            # persist new project under departamento
            PROJECTS_BY_DEPT.setdefault(departamento, [])
            if proyecto not in PROJECTS_BY_DEPT[departamento]:
                PROJECTS_BY_DEPT[departamento].append(proyecto)
                save_projects()
        elif proyecto_especial:
            proyecto = proyecto_especial
            PROJECTS_BY_DEPT.setdefault(departamento, [])
            if proyecto not in PROJECTS_BY_DEPT[departamento]:
                PROJECTS_BY_DEPT[departamento].append(proyecto)
                save_projects()

        ubicacion=request.form.get("ubicacion")
        motivo=request.form.get("motivo","")
        lines_text=request.form.get("products_json","").strip()
        if not lines_text:
            flash("Agrega al menos un producto","danger")
            return redirect(url_for("solicitud"))
        master_id=int(datetime.datetime.now().timestamp())
        wb=openpyxl.load_workbook(REQUISITIONS_FILE)
        ws=wb["Requisiciones"]
        jefe_key = USERS.get(session["user"], {}).get("jefe", "")
        jefe_name = USERS.get(jefe_key, {}).get("name", jefe_key if jefe_key else "")
        lines=[l.strip() for l in lines_text.splitlines() if l.strip()]
        for idx,line in enumerate(lines):
            parts=[p.strip() for p in line.split("|")]
            producto=parts[0] if len(parts)>0 else ""
            unidad=parts[1] if len(parts)>1 else ""
            cantidad=parts[2] if len(parts)>2 else ""
            image_filename=""
            file_field=f"image_{idx}"
            if file_field in request.files:
                f=request.files[file_field]
                if f and f.filename and allowed_file(f.filename):
                    filename=secure_filename(f.filename)
                    unique=f"{uuid.uuid4().hex}_{filename}"
                    path=os.path.join(UPLOAD_FOLDER,unique)
                    f.save(path)
                    image_filename=unique
                    try:
                        ws_img=wb["Images"] if "Images" in wb.sheetnames else wb.create_sheet("Images")
                        ws_img.append([master_id,unique])
                        try:
                            img=XLImage(path)
                            ws_img.add_image(img,f"A{ws_img.max_row}")
                        except:
                            pass
                    except:
                        pass
            ws.append([master_id,datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                       session["user"],session["name"],departamento,etapa,proyecto,
                       producto,unidad,cantidad,ubicacion,motivo,jefe_name,
                       "Pendiente","","",image_filename])
        wb.save(REQUISITIONS_FILE)
        flash("Requisición enviada correctamente","success")
        return redirect(url_for("solicitud"))
    return render_template("solicitud.html", products=products, departments=DEPARTMENTS,
                           stages=STAGES, projects_by_dept=PROJECTS_BY_DEPT, units=UNITS, locations=LOCATIONS, USERS=USERS)

@app.route("/mis_requisiciones", methods=["GET","POST"])
def mis_requisiciones():
    if "user" not in session:
        return redirect(url_for("login"))
    # POST para actualizar ID (administrador)
    if request.method=="POST" and session.get("role")=="admin":
        if request.form.get("action")=="update_id":
            old_id=request.form.get("old_id")
            new_id=request.form.get("new_id")
            if old_id and new_id:
                try:
                    old_id_val=int(old_id)
                    new_id_val=int(new_id)
                    wb=openpyxl.load_workbook(REQUISITIONS_FILE)
                    ws=wb["Requisiciones"]
                    changed=0
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value==old_id_val:
                            row[0].value=new_id_val
                            changed+=1
                    wb.save(REQUISITIONS_FILE)
                    flash(f"ID actualizado en {changed} fila(s).","success")
                except:
                    flash("Los IDs deben ser numéricos","danger")
        return redirect(url_for("mis_requisiciones"))
    wb=openpyxl.load_workbook(REQUISITIONS_FILE,data_only=True)
    ws=wb["Requisiciones"]
    rows=[]
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Mostrar si admin o si es del usuario o es su jefe (por nombre)
        if session.get("role")=="admin" or row[2]==session.get("user") or row[12]==USERS.get(session.get("user"),{}).get("name"):
            rows.append(row)
    return render_template("mis_requisiciones.html", requisiciones=rows)

@app.route("/requisiciones/download")
def download_requisiciones():
    # Alias: solo admin can download all (redirect to admin download route)
    return redirect(url_for("download_and_clear_requisitions"))

@app.route("/descargar_requisiciones")
def alias_descargar_requisiciones():
    return redirect(url_for("download_and_clear_requisitions"))

@app.route("/admin/products", methods=["GET","POST"])
def admin_products():
    if "user" not in session or session.get("role")!="admin":
        return "Acceso denegado", 403
    products=get_products()
    if request.method=="POST":
        action=request.form.get("action")
        name=request.form.get("name","").strip()
        if action=="add" and name:
            if name not in products:
                products.append(name)
                save_products(products)
                flash("Producto agregado","success")
            else:
                flash("El producto ya existe","warning")
        elif action=="update":
            old=request.form.get("old_name","").strip()
            new=request.form.get("new_name","").strip()
            if old and new and old in products:
                idx=products.index(old)
                products[idx]=new
                save_products(products)
                flash("Producto actualizado","success")
            else:
                flash("Producto no encontrado","danger")
        elif action=="delete":
            if name in products:
                products=[p for p in products if p!=name]
                save_products(products)
                flash("Producto eliminado","success")
            else:
                flash("Producto no encontrado","danger")
        return redirect(url_for("admin_products"))
    return render_template("admin_products.html", products=products)

@app.route("/api/products")
def api_products():
    q=request.args.get("q","").lower()
    products=get_products()
    return jsonify([p for p in products if q in p.lower()][:200])

@app.route("/api/projects/<dept>")
def api_projects(dept):
    load_projects()
    return jsonify(PROJECTS_BY_DEPT.get(dept,[]))

@app.route("/api/stages")
def api_stages():
    return jsonify(STAGES)

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route("/admin/upload_image", methods=["POST"])
def admin_upload_image():
    if "user" not in session or session.get("role")!="admin":
        return "Acceso denegado", 403
    req_id = request.form.get("req_id")
    file = request.files.get("file")
    if not req_id or not file or file.filename=="":
        flash("Falta ID o archivo", "danger")
        return redirect(url_for("mis_requisiciones"))
    if not allowed_file(file.filename):
        flash("Tipo de archivo no permitido. Solo imágenes.", "danger")
        return redirect(url_for("mis_requisiciones"))
    try:
        wb = openpyxl.load_workbook(REQUISITIONS_FILE)
        ws = wb["Requisiciones"]
        req_id_int = None
        try:
            req_id_int = int(req_id)
        except:
            req_id_int = None
        found=False
        for row in ws.iter_rows(min_row=2):
            cell_id = row[0].value
            if (req_id_int is not None and cell_id==req_id_int) or (str(cell_id)==req_id):
                filename = secure_filename(file.filename)
                unique = f"{uuid.uuid4().hex}_{filename}"
                path = os.path.join(UPLOAD_FOLDER, unique)
                file.save(path)
                oldval = row[16].value if len(row)>16 else None
                if oldval:
                    oldpath = os.path.join(UPLOAD_FOLDER, str(oldval))
                    try:
                        if os.path.exists(oldpath): os.remove(oldpath)
                    except:
                        pass
                if len(row)>16:
                    row[16].value = unique
                else:
                    ws.cell(row=row[0].row, column=17, value=unique)
                ws_img = wb["Images"] if "Images" in wb.sheetnames else wb.create_sheet("Images")
                ws_img.append([cell_id, unique])
                try:
                    img = XLImage(path)
                    ws_img.add_image(img, f"A{ws_img.max_row}")
                except:
                    pass
                found=True
                break
        if found:
            wb.save(REQUISITIONS_FILE)
            flash("Imagen subida y asociada a la requisición.", "success")
        else:
            flash("ID de requisición no encontrado.", "danger")
    except Exception as e:
        print("Error al subir imagen:", e)
        flash("Error al procesar la imagen.", "danger")
    return redirect(url_for("mis_requisiciones"))

@app.route("/admin/download_requisition/<req_id>")
def admin_download_requisition(req_id):
    if "user" not in session or session.get("role")!="admin":
        return "Acceso denegado", 403
    wb = openpyxl.load_workbook(REQUISITIONS_FILE,data_only=True)
    ws = wb["Requisiciones"]
    rows_to_write=[]
    headers = [cell for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (str(row[0])==str(req_id)) or (isinstance(row[0], int) and str(row[0])==req_id):
            rows_to_write.append(list(row))
    if not rows_to_write:
        flash("No se encontró la requisición.", "danger")
        return redirect(url_for("mis_requisiciones"))
    outname = f"requisicion_{req_id}.csv"
    with open(outname,"w",newline="",encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows_to_write:
            writer.writerow(r)
    return send_file(outname, as_attachment=True)

# --------------------------
# AUTORIZACIONES (Solo Jefes y Admin)
# --------------------------
@app.route("/autorizaciones", methods=["GET", "POST"])
def autorizaciones():
    if "user" not in session:
        return redirect(url_for("login"))

    user_role = session.get("role")
    user_name = USERS.get(session.get("user"), {}).get("name", "")

    # Determinar si es jefe (actual: si algún usuario tiene su 'jefe' igual a este username)
    is_jefe = False
    for k,v in USERS.items():
        if v.get('jefe') == session.get('user'):
            is_jefe = True
            break
    if user_role not in ["admin"] and not is_jefe:
        flash("Acceso denegado: esta sección es solo para jefes o administradores.", "danger")
        return redirect(url_for("solicitud"))

    wb = openpyxl.load_workbook(REQUISITIONS_FILE)
    ws = wb["Requisiciones"]

    if request.method == "POST":
        req_id = request.form.get("req_id")
        decision = request.form.get("decision")
        comentario = request.form.get("comentario", "")
        updated = False
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(req_id):
                if decision == "aprobar":
                    row[13].value = "Aprobada"
                elif decision == "rechazar":
                    row[13].value = "Rechazada"
                row[14].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                row[15].value = comentario
                updated = True
                break
        if updated:
            wb.save(REQUISITIONS_FILE)
            flash("Estatus actualizado correctamente.", "success")
        else:
            flash("No se encontró la requisición.", "danger")
        return redirect(url_for("autorizaciones"))

    # Mostrar requisiciones pendientes del jefe o admin
    rows = []
    wb_ro = openpyxl.load_workbook(REQUISITIONS_FILE, data_only=True)
    ws_ro = wb_ro["Requisiciones"]
    for row in ws_ro.iter_rows(min_row=2, values_only=True):
        # admin ve todo, jefe ve los que tengan como Jefe su nombre
        if user_role == "admin" or row[12] == user_name:
            if row[13] == "Pendiente":
                rows.append(row)
    return render_template("autorizaciones.html", requisiciones=rows)

# --------------------------
# Descargar todas las requisiciones como Excel y limpiar Requisiciones (solo admin)
# --------------------------
@app.route("/download_and_clear_requisitions")
def download_and_clear_requisitions():
    if "user" not in session or session.get("role")!="admin":
        flash("Acceso denegado.", "danger")
        return redirect(url_for("login"))

    # Leer todas las requisiciones actuales
    wb = openpyxl.load_workbook(REQUISITIONS_FILE, data_only=True)
    ws = wb["Requisiciones"]
    headers = [c for c in ws[1]]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

    if not rows:
        flash("No hay requisiciones para descargar.", "info")
        return redirect(url_for("autorizaciones"))

    # Crear DataFrame y guardarlo en static/descargas
    try:
        df = pd.DataFrame(rows, columns=headers)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        outpath = os.path.join(DOWNLOADS_FOLDER, f"solicitudes_{timestamp}.xlsx")
        df.to_excel(outpath, index=False)
    except Exception as e:
        print("Error creando Excel:", e)
        flash("Error generando el archivo Excel.", "danger")
        return redirect(url_for("autorizaciones"))

    # Ahora limpiar la hoja Requisiciones (dejar solo la fila de encabezado)
    try:
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Requisiciones"
        ws2.append(headers)
        # Also recreate Images sheet
        ws_img = wb2.create_sheet("Images")
        ws_img.append(["ID", "ImageFile"])
        wb2.save(REQUISITIONS_FILE)
    except Exception as e:
        print("Error limpiando Requisiciones:", e)
        flash("Error al limpiar las requisiciones.", "danger")
        return redirect(url_for("autorizaciones"))

    flash(f"Archivo guardado en: static/descargas/{os.path.basename(outpath)}. Requisiciones limpiadas.", "success")
    # Ofrecer descarga al admin
    return send_file(outpath, as_attachment=True)



# --- Compradores: ver y actualizar estatus ---
def read_reqs_df():
    # Reads the Excel into a dataframe and returns it
    path = os.path.join(os.path.dirname(__file__),'Requisiciones.xlsx')
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, engine='openpyxl')
    except Exception:
        df = pd.DataFrame()
    return df

def write_reqs_df(df):
    path = os.path.join(os.path.dirname(__file__),'Requisiciones.xlsx')
    df.to_excel(path, index=False)

@app.route('/compras', methods=['GET','POST'])
def compras():
    if "user" not in session:
        return redirect(url_for('login'))
    if session.get("role") not in ['comprador','admin']:
        flash("Acceso denegado: solo compradores o administradores","danger")
        return redirect(url_for('solicitud'))
    df = read_reqs_df()
    if df.empty:
        requis = []
    else:
        # Filter relevant statuses for compras
        requis = df[df['Estatus'].isin(['En cotizacion','En espera de entrega','Por comprar','Pendiente'])].to_dict(orient='records')
    if request.method=='POST':
        idx = int(request.form.get('idx'))
        nuevo = request.form.get('nuevo_estatus')
        if not df.empty and 0 <= idx < len(df):
            df.at[idx,'Estatus'] = nuevo
            write_reqs_df(df)
            flash("Estatus actualizado","success")
            return redirect(url_for('compras'))
    return render_template('compras.html', requisiciones=requis)

if __name__=="__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
