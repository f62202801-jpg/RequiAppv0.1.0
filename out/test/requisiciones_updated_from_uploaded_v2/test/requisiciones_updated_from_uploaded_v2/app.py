from flask import Flask, render_template, request, redirect, session, url_for, send_from_directory, jsonify, flash
import os, json, datetime, uuid
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get("APP_SECRET", "CAMBIAR_POR_SECRETO_PROD")

BASE_DIR = os.path.dirname(__file__)
CATALOG_FILE = os.path.join(BASE_DIR, "Control De Herramientas.xlsm")
REQUISITIONS_FILE = os.path.join(BASE_DIR, "Requisiciones.xlsx")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXT = {"png", "jpg", "jpeg", "gif"}

USERS = {
    "erik": {"password": "1234", "role": "solicitante", "name": "Erik Oronzor", "jefe": "jefe"},
    "juan": {"password": "juan123", "role": "solicitante", "name": "Juan Perez", "jefe": "jefe"},
    "jefe": {"password": "admin", "role": "jefe", "name": "Jefe Inmediato"},
    "admin": {"password": "admin", "role": "admin", "name": "Administrador"}
}

DEPARTMENTS = ["Fresa", "Invernadero", "Administracion"]

STAGES = {
    "Fresa": ["Preplantacion", "Plantacion", "Crecimiento", "Cumplimiento de pedido", "Cosecha", "Postcosecha", "Mercancias en transito"],
    "Invernadero": ["Preplantacion", "Plantacion", "Crecimiento", "Cumplimiento de pedido", "Cosecha/Postcosecha", "Mercancias en transito"],
    "Administracion": ["Gastos de administracion", "Opex", "Capex"]
}

PROJECTS_BY_DEPT = {
    "Fresa": ["101 - Strawberry"],
    "Invernadero": ["103 - Raspberry Increace", "106 - Raspberry Puebla", "110 - Blackberry Increce", "111 - Blueberry Puebla",
                    "207 - Blackberry Puebla", "210 - Crop Rotation Oats", "211 - Crop Rotation Barley", "212 - TNP Plantfor",
                    "213 - TNP NAP", "216 - Raspberry Queretaro", "217 - Blackberry Queretaro", "218 - Blueberry Queretaro"],
    "Administracion": ["300 - Board Of Directors, President & Special Projects", "301 - Sales", "302 - Outdoor T&E Production",
                       "303 - Indoor T&E Production", "304 - Finance & Controlling", "305 - HR , SSHE and Facility",
                       "306 - Procurement & Warehouse", "307- Legal", "308 - IT"]
}

UNITS = ["Bulto", "Gramos", "Juego", "Kilogramo", "Litro", "Metro", "Mililitros", "Par", "Pieza", "Rollo", "Paquete", "Tonelada"]

LOCATIONS = ["Almacén", "Baja California", "Cowork - Puebla", "Cowork - Guadalajara", "Ocotepec", "Praderas", "Purísima de Cubos",
             "San Martín", "San Roque", "Santa Julia 1", "Santa Julia 2", "Santa Julia 3", "Santa Julia 4", "Santa Julia 5",
             "Santa Julia 6", "Santa Lugarda", "Zamora"]

TOKENS_FILE = os.path.join(BASE_DIR, "device_tokens.json")

def init_files():
    if not os.path.exists(TOKENS_FILE):
        with open(TOKENS_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)
    if not os.path.exists(REQUISITIONS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requisiciones"
        ws.append(["ID", "Fecha", "Usuario", "Solicitante", "Departamento", "Etapa", "Proyecto",
                   "Productos_JSON", "Unidad_JSON", "Cantidad_JSON", "Ubicacion", "Motivo",
                   "Jefe Inmediato", "Estatus", "Fecha Autorizacion", "Comentarios Jefe", "ImageFiles"])
        ws2 = wb.create_sheet("Images")
        ws2.append(["ID", "ImageFile"])
        wb.save(REQUISITIONS_FILE)

init_files()

def get_products_from_catalog():
    products = []
    if os.path.exists(CATALOG_FILE):
        try:
            wb = openpyxl.load_workbook(CATALOG_FILE, data_only=True)
            if "PRODUCTOS" in wb.sheetnames:
                ws = wb["PRODUCTOS"]
                for cell in ws['A']:
                    if cell.value and str(cell.value).strip() != '':
                        products.append(str(cell.value).strip())
        except Exception as e:
            print("Error reading catalog:", e)
    return products

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form.get("usuario")
        pwd = request.form.get("password")
        u = USERS.get(user)
        if u and u["password"] == pwd:
            session["user"] = user
            session["role"] = u["role"]
            session["name"] = u["name"]
            return redirect(url_for("solicitud"))
        flash("Usuario o contraseña incorrectos", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/solicitud", methods=["GET", "POST"])
def solicitud():
    if "user" not in session:
        return redirect(url_for("login"))
    products_catalog = get_products_from_catalog()
    if request.method == "POST":
        departamento = request.form.get("departamento")
        etapa = request.form.get("etapa")
        proyecto = request.form.get("proyecto")
        ubicacion = request.form.get("ubicacion")
        motivo = request.form.get("motivo", "")
        products_json = request.form.get("products_json")
        units_json = request.form.get("units_json")
        qty_json = request.form.get("qty_json")
        try:
            products_list = json.loads(products_json)
            units_list = json.loads(units_json)
            qty_list = json.loads(qty_json)
        except:
            return "Datos de productos inválidos", 400

        wb = openpyxl.load_workbook(REQUISITIONS_FILE)
        ws = wb["Requisiciones"]
        next_id = ws.max_row

        # 🔹 Corregido: guardar el nombre del jefe inmediato, no solo el username
        jefe_username = USERS.get(session["user"], {}).get("jefe", "")
        jefe_name = USERS.get(jefe_username, {}).get("name", "")

        image_files = []
        for idx in range(len(products_list)):
            file_field = f"image_{idx}"
            if file_field in request.files:
                f = request.files[file_field]
                if f and f.filename and allowed_file(f.filename):
                    filename = secure_filename(f.filename)
                    unique = f"{uuid.uuid4().hex}_{filename}"
                    path = os.path.join(UPLOAD_FOLDER, unique)
                    f.save(path)
                    image_files.append(unique)
                    try:
                        ws_img = wb["Images"] if "Images" in wb.sheetnames else wb.create_sheet("Images")
                        ws_img.append([next_id, unique])
                        img = XLImage(path)
                        anchor = f"A{ws_img.max_row}"
                        ws_img.add_image(img, anchor)
                    except Exception as e:
                        print("Embed image error:", e)

        ws.append([next_id,
                   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                   session["user"], session["name"], departamento, etapa, proyecto,
                   json.dumps(products_list, ensure_ascii=False),
                   json.dumps(units_list, ensure_ascii=False),
                   json.dumps(qty_list, ensure_ascii=False),
                   ubicacion, motivo, jefe_name,
                   "Pendiente", "", "", ";".join(image_files)])
        wb.save(REQUISITIONS_FILE)
        flash("Requisición enviada correctamente", "success")
        return redirect(url_for("solicitud"))
    return render_template("solicitud.html", products=products_catalog, departments=DEPARTMENTS,
                           stages=STAGES, projects_by_dept=PROJECTS_BY_DEPT, units=UNITS, locations=LOCATIONS)

@app.route("/mis_requisiciones")
def mis_requisiciones():
    if "user" not in session:
        return redirect(url_for("login"))
    wb = openpyxl.load_workbook(REQUISITIONS_FILE, data_only=True)
    ws = wb["Requisiciones"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if session.get("role") == "admin" or row[2] == session.get("user") or row[12] == USERS.get(session.get("user"), {}).get("name"):
            rows.append(row)
    return render_template("mis_requisiciones.html", requisiciones=rows)

@app.route("/autorizar", methods=["GET", "POST"])
def autorizar():
    if "user" not in session or session.get("role") not in ("jefe", "admin"):
        return "Acceso denegado", 403
    wb = openpyxl.load_workbook(REQUISITIONS_FILE)
    ws = wb["Requisiciones"]
    pending = []
    for row in ws.iter_rows(min_row=2):
        status = row[13].value
        jefe = row[12].value
        if status == "Pendiente" and (session.get("role") == "admin" or jefe == session.get("name")):
            pending.append([cell.value for cell in row])
    if request.method == "POST":
        req_id = int(request.form.get("req_id"))
        accion = request.form.get("accion")
        comentarios = request.form.get("comentarios", "")
        for row in ws.iter_rows(min_row=2):
            if row[0].value == req_id:
                if accion == "autorizar":
                    row[13].value = "Aprobado"
                    row[14].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row[15].value = comentarios
                else:
                    row[13].value = "Rechazado"
                    row[14].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row[15].value = comentarios
                break
        wb.save(REQUISITIONS_FILE)
        flash("Acción registrada", "success")
        return redirect(url_for("autorizar"))
    return render_template("autorizar.html", pendientes=pending)

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route("/api/projects/<dept>")
def api_projects(dept):
    return jsonify(PROJECTS_BY_DEPT.get(dept, []))

@app.route("/api/stages")
def api_stages():
    return jsonify(STAGES)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
